from pathlib import Path
import asyncio, json, shutil, re, io, threading, secrets
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from fastapi import FastAPI, Form, Request, UploadFile, File, Depends, HTTPException
from fastapi.responses import (
    HTMLResponse, JSONResponse, RedirectResponse,
    FileResponse, StreamingResponse, Response,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import or_, func

from .database import Base, engine, get_db, SessionLocal
from .models import DocumentRecord, User
from .auth import (
    hash_password, verify_password, validate_password,
    UPLOAD_ROLES, EDIT_ROLES, SEARCH_ROLES, DASHBOARD_ROLES, ADMIN_ROLES,
)
from .services.pdf_splitter import split_pdf
from .services.ocr_service import ocr_pdf_page, ocr_structured_fields
from .services.extractor import extract_fields, extract_fields_from_structured, extract_fields_native
from .services.file_namer import sanitize_filename

# ── directorios ───────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parent.parent
APP_DIR    = BASE_DIR / "app"
DATA_DIR   = BASE_DIR / "data"
INPUT_DIR  = DATA_DIR / "input"
OUTPUT_DIR = DATA_DIR / "output"

INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Clave de sesión persistente
_KEY_FILE = DATA_DIR / ".session_key"
if not _KEY_FILE.exists():
    _KEY_FILE.write_text(secrets.token_hex(32))
SESSION_SECRET = _KEY_FILE.read_text().strip()

Base.metadata.create_all(bind=engine)

# ── app ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="ISSEMYM OCR App")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, max_age=28800, https_only=False)
app.mount("/static", StaticFiles(directory=str(APP_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))

# Helper disponible en todos los templates: {% if has_role(current_user, "admin", "adjuntar") %}
templates.env.globals["has_role"] = lambda user, *roles: user.has_any(*roles)

_file_lock = threading.Lock()


# ── admin inicial ─────────────────────────────────────────────────────────────
@app.on_event("startup")
def create_initial_admin():
    db = SessionLocal()
    try:
        if not db.query(User).first():
            admin = User(
                username="admin",
                email="admin@sistema.local",
                full_name="Administrador",
                hashed_password=hash_password("Admin@2026!"),
                role="admin",
                is_active=True,
            )
            db.add(admin)
            db.commit()
            print("\n" + "="*55)
            print("  USUARIO ADMINISTRADOR CREADO")
            print("  Usuario:    admin")
            print("  Contraseña: Admin@2026!")
            print("  Cambia la contraseña al primer inicio de sesión.")
            print("="*55 + "\n")
    finally:
        db.close()



# ── helpers de autenticación ──────────────────────────────────────────────────
def _get_user(request: Request, db: Session):
    uid = request.session.get("user_id")
    if not uid:
        return None
    return db.query(User).filter(User.id == int(uid), User.is_active == True).first()


def _auth(request: Request, db: Session, *roles: str):
    """Devuelve (user, None) si OK, o (None/user, response) si hay que redirigir/403."""
    uid = request.session.get("user_id")
    if not uid:
        return None, RedirectResponse("/login", status_code=302)
    user = db.query(User).filter(User.id == int(uid), User.is_active == True).first()
    if not user:
        request.session.clear()
        return None, RedirectResponse("/login", status_code=302)
    if roles and not user.has_any(*roles):
        return user, templates.TemplateResponse(
            "403.html", {"request": request, "current_user": user}, status_code=403
        )
    return user, None


# ── login / logout ────────────────────────────────────────────────────────────
@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request):
    if request.session.get("user_id"):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login", response_class=HTMLResponse)
def login_post(
    request: Request,
    username: str = Form(""),
    password: str = Form(""),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.username == username.strip()).first()
    if not user or not user.is_active or not verify_password(password, user.hashed_password):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Usuario o contraseña incorrectos.", "username": username},
            status_code=401,
        )
    request.session["user_id"] = str(user.id)
    return RedirectResponse("/", status_code=302)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)


# ── perfil / cambio de contraseña ─────────────────────────────────────────────
@app.get("/perfil", response_class=HTMLResponse)
def perfil_get(request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db)
    if err:
        return err
    return templates.TemplateResponse("perfil.html", {"request": request, "current_user": user})


@app.post("/perfil", response_class=HTMLResponse)
def perfil_post(
    request: Request,
    db: Session = Depends(get_db),
    current_password: str = Form(""),
    new_password: str = Form(""),
    confirm_password: str = Form(""),
):
    user, err = _auth(request, db)
    if err:
        return err

    errors = []
    if not verify_password(current_password, user.hashed_password):
        errors.append("La contraseña actual es incorrecta.")
    if new_password != confirm_password:
        errors.append("Las contraseñas nuevas no coinciden.")
    errors += validate_password(new_password)

    if errors:
        return templates.TemplateResponse(
            "perfil.html", {"request": request, "current_user": user, "errors": errors}
        )

    user.hashed_password = hash_password(new_password)
    db.commit()
    return templates.TemplateResponse(
        "perfil.html",
        {"request": request, "current_user": user, "message": "Contraseña actualizada correctamente."},
    )


# ── admin: gestión de usuarios ────────────────────────────────────────────────
@app.get("/admin/usuarios", response_class=HTMLResponse)
def admin_usuarios(request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db, *ADMIN_ROLES)
    if err:
        return err
    users = db.query(User).order_by(User.id).all()
    msg = request.session.pop("flash", None)
    return templates.TemplateResponse(
        "admin_usuarios.html",
        {"request": request, "current_user": user, "users": users, "message": msg},
    )


@app.get("/admin/usuarios/nuevo", response_class=HTMLResponse)
def admin_nuevo_get(request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db, *ADMIN_ROLES)
    if err:
        return err
    return templates.TemplateResponse(
        "admin_usuario_form.html",
        {"request": request, "current_user": user, "edit_user": None, "errors": [], "form": {}},
    )


@app.post("/admin/usuarios/nuevo", response_class=HTMLResponse)
async def admin_nuevo_post(request: Request, db: Session = Depends(get_db)):
    current_user, err = _auth(request, db, *ADMIN_ROLES)
    if err:
        return err

    fd = await request.form()
    username      = fd.get("username", "").strip()
    full_name     = fd.get("full_name", "").strip()
    email         = fd.get("email", "").strip() or None
    roles_list    = fd.getlist("roles")
    password      = fd.get("password", "")
    password_confirm = fd.get("password_confirm", "")

    _VALID_ROLES = {"busqueda", "adjuntar", "dashboards", "admin"}
    roles_list = [r for r in roles_list if r in _VALID_ROLES]
    role_str = ",".join(sorted(roles_list))
    form_data = {"username": username, "full_name": full_name, "email": email, "roles": roles_list}
    errors = []

    if not username:
        errors.append("El nombre de usuario es obligatorio.")
    elif db.query(User).filter(User.username == username).first():
        errors.append(f'El usuario "{username}" ya existe.')
    if not full_name:
        errors.append("El nombre completo es obligatorio.")
    if not roles_list:
        errors.append("Asigna al menos un perfil al usuario.")
    if password != password_confirm:
        errors.append("Las contraseñas no coinciden.")
    errors += validate_password(password)

    if errors:
        return templates.TemplateResponse(
            "admin_usuario_form.html",
            {"request": request, "current_user": current_user,
             "edit_user": None, "errors": errors, "form": form_data},
        )

    db.add(User(
        username=username, full_name=full_name, email=email,
        hashed_password=hash_password(password),
        role=role_str, is_active=True, created_by_id=current_user.id,
    ))
    db.commit()
    request.session["flash"] = f'Usuario "{username}" creado correctamente.'
    return RedirectResponse("/admin/usuarios", status_code=302)


@app.get("/admin/usuarios/{uid}/editar", response_class=HTMLResponse)
def admin_editar_get(uid: int, request: Request, db: Session = Depends(get_db)):
    current_user, err = _auth(request, db, *ADMIN_ROLES)
    if err:
        return err
    edit_user = db.query(User).filter(User.id == uid).first()
    if not edit_user:
        raise HTTPException(404, "Usuario no encontrado")
    return templates.TemplateResponse(
        "admin_usuario_form.html",
        {"request": request, "current_user": current_user, "edit_user": edit_user, "errors": [], "form": {}},
    )


@app.post("/admin/usuarios/{uid}/editar", response_class=HTMLResponse)
async def admin_editar_post(uid: int, request: Request, db: Session = Depends(get_db)):
    current_user, err = _auth(request, db, *ADMIN_ROLES)
    if err:
        return err

    edit_user = db.query(User).filter(User.id == uid).first()
    if not edit_user:
        raise HTTPException(404, "Usuario no encontrado")

    fd = await request.form()
    full_name        = fd.get("full_name", "").strip()
    email            = fd.get("email", "").strip() or None
    roles_list       = fd.getlist("roles")
    password         = fd.get("password", "")
    password_confirm = fd.get("password_confirm", "")

    _VALID_ROLES = {"busqueda", "adjuntar", "dashboards", "admin"}
    roles_list = [r for r in roles_list if r in _VALID_ROLES]
    role_str = ",".join(sorted(roles_list))
    form_data = {"full_name": full_name, "email": email, "roles": roles_list}
    errors = []

    if not full_name:
        errors.append("El nombre completo es obligatorio.")
    if not roles_list:
        errors.append("Asigna al menos un perfil al usuario.")
    if password:
        if password != password_confirm:
            errors.append("Las contraseñas no coinciden.")
        errors += validate_password(password)

    if errors:
        return templates.TemplateResponse(
            "admin_usuario_form.html",
            {"request": request, "current_user": current_user,
             "edit_user": edit_user, "errors": errors, "form": form_data},
        )

    edit_user.full_name = full_name
    edit_user.email = email
    edit_user.role = role_str
    if password:
        edit_user.hashed_password = hash_password(password)
    db.commit()
    request.session["flash"] = f'Usuario "{edit_user.username}" actualizado correctamente.'
    return RedirectResponse("/admin/usuarios", status_code=302)


@app.post("/admin/usuarios/{uid}/toggle")
def admin_toggle(uid: int, request: Request, db: Session = Depends(get_db)):
    current_user, err = _auth(request, db, *ADMIN_ROLES)
    if err:
        return err
    edit_user = db.query(User).filter(User.id == uid).first()
    if not edit_user:
        raise HTTPException(404, "Usuario no encontrado")
    if edit_user.id == current_user.id:
        raise HTTPException(400, "No puedes desactivar tu propia cuenta")
    edit_user.is_active = not edit_user.is_active
    db.commit()
    estado = "activado" if edit_user.is_active else "desactivado"
    request.session["flash"] = f'Usuario "{edit_user.username}" {estado}.'
    return RedirectResponse("/admin/usuarios", status_code=302)


# ── portal ────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db)
    if err:
        return err
    return templates.TemplateResponse("index.html", {"request": request, "current_user": user})


def _event(data: dict) -> str:
    return f"data: {json.dumps(data)}\n\n"


# ── procesamiento de PDF ──────────────────────────────────────────────────────
def _extract_and_rename(single_pdf: Path, idx: int, total: int, loop, queue):
    loop.call_soon_threadsafe(queue.put_nowait, {
        "step": "extract", "current": idx, "total": total,
        "msg": f"Extrayendo información ({idx} de {total})"
    })
    fields = extract_fields_native(single_pdf)
    if fields is None:
        structured_raw = ocr_structured_fields(single_pdf, 0)
        fields = extract_fields_from_structured(structured_raw)
        if not fields.get("rfc") and not fields.get("tipo_movimiento"):
            full_text = ocr_pdf_page(single_pdf, 0)
            fields = extract_fields(full_text)

    clave = fields.get("clave_afiliacion_issemym")
    base_name = f"{sanitize_filename(clave)}.pdf" if clave else f"SIN_CLAVE_{idx:04d}.pdf"
    with _file_lock:
        final_path = single_pdf.with_name(base_name)
        counter = 1
        while final_path.exists() and final_path.name != single_pdf.name:
            stem = final_path.stem
            if "_dup" in stem:
                stem = stem.rsplit("_dup", 1)[0]
            final_path = final_path.with_name(f"{stem}_dup{counter}.pdf")
            counter += 1
        single_pdf.rename(final_path)
    return fields, final_path, idx


@app.post("/process")
async def process_pdf(request: Request, pdf_file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user or not user.has_any(*UPLOAD_ROLES):
        return JSONResponse({"error": "Sin autorización"}, status_code=403)

    content = await pdf_file.read()
    filename = pdf_file.filename
    loop = asyncio.get_event_loop()
    queue: asyncio.Queue = asyncio.Queue()

    def process_sync():
        db2 = SessionLocal()
        try:
            loop.call_soon_threadsafe(queue.put_nowait, {"step": "upload", "msg": f"Guardando archivo: {filename}"})
            destination = INPUT_DIR / filename
            destination.write_bytes(content)

            loop.call_soon_threadsafe(queue.put_nowait, {"step": "split", "msg": "Dividiendo páginas..."})
            split_files = split_pdf(destination, OUTPUT_DIR)
            total = len(split_files)
            loop.call_soon_threadsafe(queue.put_nowait, {"step": "split_done", "msg": f"Dividido en {total} páginas", "total": total})

            max_workers = min(4, total)
            page_results = []
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(_extract_and_rename, pdf, idx, total, loop, queue): idx
                    for idx, pdf in enumerate(split_files, start=1)
                }
                done_count = 0
                for future in as_completed(futures):
                    fields, final_path, idx = future.result()
                    done_count += 1
                    loop.call_soon_threadsafe(queue.put_nowait, {
                        "step": "rename", "current": done_count, "total": total,
                        "msg": f"Guardando registro {done_count} de {total}"
                    })
                    page_results.append((fields, final_path, idx))

            page_results.sort(key=lambda x: x[2])
            processed = skipped = 0
            for fields, final_path, idx in page_results:
                clave = fields.get("clave_afiliacion_issemym")
                a_partir = fields.get("a_partir_de")
                tipo = fields.get("tipo_movimiento")
                if clave:
                    dup = db2.query(DocumentRecord).filter(
                        DocumentRecord.clave_afiliacion_issemym == clave,
                        DocumentRecord.a_partir_de == a_partir,
                        DocumentRecord.tipo_movimiento == tipo,
                    ).first()
                    if dup:
                        skipped += 1
                        continue
                db2.add(DocumentRecord(
                    tipo_movimiento=tipo, a_partir_de=a_partir,
                    clave_afiliacion_issemym=clave, rfc=fields.get("rfc"),
                    nombre_completo=fields.get("nombre_completo"),
                    institucion_publica=fields.get("institucion_publica"),
                    clave_institucion_publica=fields.get("clave_institucion_publica"),
                    nombramiento_categoria=fields.get("nombramiento_categoria"),
                    fecha_emision=fields.get("fecha_emision"),
                    firma_cadena_digital=fields.get("firma_cadena_digital"),
                    nombre_archivo_pdf=final_path.name,
                    ruta_archivo_pdf=str(final_path),
                    pagina_origen=idx,
                    texto_extraido=fields.get("texto_extraido"),
                ))
                processed += 1

            db2.commit()
            msg = f"Proceso completado. {processed} documentos guardados."
            if skipped:
                msg += f" ({skipped} duplicados omitidos)"
            loop.call_soon_threadsafe(queue.put_nowait, {"step": "done", "processed": processed, "msg": msg})
        except Exception as e:
            loop.call_soon_threadsafe(queue.put_nowait, {"step": "error", "msg": str(e)})
        finally:
            db2.close()
            loop.call_soon_threadsafe(queue.put_nowait, None)

    asyncio.get_event_loop().run_in_executor(None, process_sync)

    async def generate():
        while True:
            item = await queue.get()
            if item is None:
                break
            yield _event(item)

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── búsqueda ──────────────────────────────────────────────────────────────────
SEARCH_FIELDS = [
    DocumentRecord.tipo_movimiento, DocumentRecord.a_partir_de,
    DocumentRecord.clave_afiliacion_issemym, DocumentRecord.rfc,
    DocumentRecord.nombre_completo, DocumentRecord.institucion_publica,
    DocumentRecord.clave_institucion_publica, DocumentRecord.nombramiento_categoria,
    DocumentRecord.fecha_emision, DocumentRecord.nombre_archivo_pdf,
    DocumentRecord.texto_extraido,
]


def _ddmmyyyy_to_yyyymmdd(s: str) -> str:
    """Convierte 'DD/MM/YYYY' → 'YYYYMMDD' para comparación lexicográfica correcta."""
    parts = s.strip().split("/")
    if len(parts) == 3:
        dd, mm, yyyy = parts[0].zfill(2), parts[1].zfill(2), parts[2].zfill(4)
        return f"{yyyy}{mm}{dd}"
    return s  # devuelve tal cual si el formato es inesperado


def _date_sortable_expr():
    """Expresión SQLAlchemy que convierte DD/MM/YYYY almacenado → YYYYMMDD para comparar.
    Usa el operador || (concatenación de texto) compatible con SQLite y PostgreSQL.
    """
    col = DocumentRecord.a_partir_de
    # func.substr devuelve un objeto genérico; usamos .op('||') para concatenación de texto
    return (
        func.substr(col, 7, 4)
        .op("||")(func.substr(col, 4, 2))
        .op("||")(func.substr(col, 1, 2))
    )


@app.get("/search", response_class=HTMLResponse)
def search(request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db, *SEARCH_ROLES)
    if err:
        return err
    return templates.TemplateResponse("search.html", {"request": request, "current_user": user})


@app.get("/api/search")
def api_search(
    request: Request,
    q: str = "", tipo: str = "", desde: str = "", hasta: str = "",
    page: int = 1, per_page: int = 5,
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user or not user.has_any(*SEARCH_ROLES):
        return JSONResponse({"error": "Sin autorización"}, status_code=403)

    query = db.query(DocumentRecord)
    for token in [t for t in q.strip().split() if t]:
        query = query.filter(or_(*[f.ilike(f"%{token}%") for f in SEARCH_FIELDS]))
    if tipo:
        tipos = [t.strip() for t in tipo.split(",") if t.strip()]
        if tipos:
            query = query.filter(DocumentRecord.tipo_movimiento.in_(tipos))
    if desde or hasta:
        date_expr = _date_sortable_expr()
        if desde:
            query = query.filter(date_expr >= _ddmmyyyy_to_yyyymmdd(desde))
        if hasta:
            query = query.filter(date_expr <= _ddmmyyyy_to_yyyymmdd(hasta))

    total = query.count()
    rows = query.order_by(DocumentRecord.id.desc()).offset((page-1)*per_page).limit(per_page).all()

    return JSONResponse({
        "total": total, "page": page, "per_page": per_page,
        "pages": max(1, -(-total // per_page)),
        "results": [{
            "id": r.id, "clave": r.clave_afiliacion_issemym or "",
            "nombre": r.nombre_completo or "", "rfc": r.rfc or "",
            "tipo": r.tipo_movimiento or "", "a_partir_de": r.a_partir_de or "",
            "fecha_emision": r.fecha_emision or "", "archivo": r.nombre_archivo_pdf or "",
        } for r in rows],
    })


@app.get("/api/stats")
def api_stats(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user or not user.has_any(*(DASHBOARD_ROLES | SEARCH_ROLES)):
        return JSONResponse({"error": "Sin autorización"}, status_code=403)

    total = db.query(func.count(DocumentRecord.id)).scalar() or 0
    by_tipo = db.query(DocumentRecord.tipo_movimiento, func.count(DocumentRecord.id))\
        .group_by(DocumentRecord.tipo_movimiento).all()
    by_inst = db.query(DocumentRecord.institucion_publica, func.count(DocumentRecord.id))\
        .group_by(DocumentRecord.institucion_publica)\
        .order_by(func.count(DocumentRecord.id).desc()).limit(10).all()
    by_year_raw = db.query(DocumentRecord.a_partir_de, func.count(DocumentRecord.id))\
        .filter(DocumentRecord.a_partir_de.isnot(None))\
        .group_by(DocumentRecord.a_partir_de).all()

    year_counts: dict = {}
    for fecha, cnt in by_year_raw:
        if fecha and len(fecha) >= 10:
            yr = fecha[-4:]
            year_counts[yr] = year_counts.get(yr, 0) + cnt

    return JSONResponse({
        "total": total,
        "by_tipo": [{"tipo": t or "Sin dato", "count": c} for t, c in by_tipo],
        "by_institucion": [{"inst": (i or "Sin dato")[:40], "count": c} for i, c in by_inst],
        "by_year": [{"year": k, "count": v} for k, v in sorted(year_counts.items())],
    })


@app.get("/api/export")
def api_export(
    request: Request,
    q: str = "", tipo: str = "", desde: str = "", hasta: str = "",
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user or not user.has_any(*SEARCH_ROLES):
        return JSONResponse({"error": "Sin autorización"}, status_code=403)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(DocumentRecord)
    for token in [t for t in q.strip().split() if t]:
        query = query.filter(or_(*[f.ilike(f"%{token}%") for f in SEARCH_FIELDS]))
    if tipo:
        tipos = [t.strip() for t in tipo.split(",") if t.strip()]
        if tipos:
            query = query.filter(DocumentRecord.tipo_movimiento.in_(tipos))
    if desde or hasta:
        date_expr = _date_sortable_expr()
        if desde:
            query = query.filter(date_expr >= _ddmmyyyy_to_yyyymmdd(desde))
        if hasta:
            query = query.filter(date_expr <= _ddmmyyyy_to_yyyymmdd(hasta))
    rows = query.order_by(DocumentRecord.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados ISSEMYM"
    burg_fill = PatternFill("solid", fgColor="8C1C40")
    alt_fill  = PatternFill("solid", fgColor="F9F0F3")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    headers = ["ID","Tipo","A partir de","Clave ISSEMYM","RFC / CURP","Nombre completo",
               "Institución pública","Clave institución","Nombramiento","Fecha emisión","Archivo PDF"]
    col_widths = [6, 10, 14, 14, 18, 36, 40, 14, 28, 14, 30]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = burg_fill
        cell.alignment = hdr_align; cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 30

    for ri, r in enumerate(rows, 2):
        vals = [r.id, r.tipo_movimiento, r.a_partir_de, r.clave_afiliacion_issemym,
                r.rfc, r.nombre_completo, r.institucion_publica, r.clave_institucion_publica,
                r.nombramiento_categoria, r.fecha_emision, r.nombre_archivo_pdf]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=val or "")
            cell.border = thin
            if ri % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ISSEMYM_export.xlsx"'},
    )


# ── dashboard ─────────────────────────────────────────────────────────────────
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db, *DASHBOARD_ROLES)
    if err:
        return err
    return templates.TemplateResponse("dashboard.html", {"request": request, "current_user": user})


# ── detalle / edición ─────────────────────────────────────────────────────────
@app.get("/detail/{record_id}", response_class=HTMLResponse)
def detail(record_id: int, request: Request, db: Session = Depends(get_db)):
    user, err = _auth(request, db, *SEARCH_ROLES)
    if err:
        return err
    record = db.query(DocumentRecord).filter(DocumentRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "Registro no encontrado")
    return templates.TemplateResponse(
        "detail.html", {"request": request, "record": record, "current_user": user}
    )


@app.post("/update/{record_id}", response_class=HTMLResponse)
def update_record(
    record_id: int, request: Request, db: Session = Depends(get_db),
    tipo_movimiento: str = Form(""), a_partir_de: str = Form(""),
    clave_afiliacion_issemym: str = Form(""), rfc: str = Form(""),
    nombre_completo: str = Form(""), institucion_publica: str = Form(""),
    clave_institucion_publica: str = Form(""), nombramiento_categoria: str = Form(""),
    fecha_emision: str = Form(""),
):
    user, err = _auth(request, db, *EDIT_ROLES)
    if err:
        return err
    record = db.query(DocumentRecord).filter(DocumentRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "Registro no encontrado")
    record.tipo_movimiento = tipo_movimiento.strip() or None
    record.a_partir_de = a_partir_de.strip() or None
    record.clave_afiliacion_issemym = clave_afiliacion_issemym.strip() or None
    record.rfc = rfc.strip() or None
    record.nombre_completo = nombre_completo.strip() or None
    record.institucion_publica = institucion_publica.strip() or None
    record.clave_institucion_publica = clave_institucion_publica.strip() or None
    record.nombramiento_categoria = nombramiento_categoria.strip() or None
    record.fecha_emision = fecha_emision.strip() or None
    db.commit()
    return templates.TemplateResponse(
        "detail.html",
        {"request": request, "record": record, "current_user": user, "message": "Cambios guardados correctamente."},
    )


@app.post("/rename/{record_id}", response_class=HTMLResponse)
def rename_record(
    record_id: int, request: Request, db: Session = Depends(get_db),
    nuevo_nombre: str = Form(""),
):
    user, err = _auth(request, db, *EDIT_ROLES)
    if err:
        return err
    record = db.query(DocumentRecord).filter(DocumentRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "Registro no encontrado")
    nuevo_nombre = nuevo_nombre.strip()
    if not nuevo_nombre:
        raise HTTPException(400, "El nombre no puede estar vacío")
    if not nuevo_nombre.lower().endswith(".pdf"):
        nuevo_nombre += ".pdf"
    old_path = Path(record.ruta_archivo_pdf)
    new_path = old_path.with_name(nuevo_nombre)
    counter = 1
    while new_path.exists() and new_path.resolve() != old_path.resolve():
        stem = nuevo_nombre.rsplit(".pdf", 1)[0]
        new_path = old_path.with_name(f"{stem}_{counter}.pdf")
        counter += 1
    if old_path.exists():
        old_path.rename(new_path)
    record.nombre_archivo_pdf = new_path.name
    record.ruta_archivo_pdf = str(new_path)
    db.commit()
    return templates.TemplateResponse(
        "detail.html",
        {"request": request, "record": record, "current_user": user, "message": f'Archivo renombrado a "{new_path.name}".'},
    )


# ── PDF / correo ──────────────────────────────────────────────────────────────
@app.get("/pdf/{record_id}")
def open_pdf(record_id: int, request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user or not user.has_any(*SEARCH_ROLES):
        return RedirectResponse("/login", status_code=302)
    record = db.query(DocumentRecord).filter(DocumentRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "Registro no encontrado")
    pdf_path = Path(record.ruta_archivo_pdf)
    if not pdf_path.exists():
        raise HTTPException(404, "Archivo PDF no encontrado")
    return FileResponse(
        path=pdf_path, filename=record.nombre_archivo_pdf, media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{record.nombre_archivo_pdf}"'},
    )


def _build_email_body(record) -> tuple[str, str]:
    """Devuelve (subject, plain_body) con los datos del registro."""
    subject = (
        f"Aviso ISSEMYM — {record.clave_afiliacion_issemym or 'SIN CLAVE'}"
        f" — {record.nombre_completo or ''}"
    )
    body = (
        f"AVISO DE MOVIMIENTO PARA LA AFILIACIÓN Y VIGENCIA DE DERECHOS\n"
        f"{'='*60}\n"
        f"Tipo de movimiento : {record.tipo_movimiento or '—'}\n"
        f"A partir de        : {record.a_partir_de or '—'}\n"
        f"Clave ISSEMYM      : {record.clave_afiliacion_issemym or '—'}\n"
        f"RFC / CURP         : {record.rfc or '—'}\n"
        f"Nombre completo    : {record.nombre_completo or '—'}\n"
        f"Institución pública: {record.institucion_publica or '—'}"
        f" ({record.clave_institucion_publica or '—'})\n"
        f"Nombramiento       : {record.nombramiento_categoria or '—'}\n"
        f"Fecha de emisión   : {record.fecha_emision or '—'}\n"
        f"{'='*60}\n"
        f"Archivo adjunto    : {record.nombre_archivo_pdf or '—'}\n"
        f"Sistema ISSEMYM Trámites Externos – Secretaría de Educación EdoMex"
    )
    return subject, body


@app.get("/api/email_data/{record_id}")
def api_email_data(record_id: int, request: Request, db: Session = Depends(get_db)):
    """Devuelve JSON con la URL de Gmail compose para componer el correo."""
    user = _get_user(request, db)
    if not user or not user.has_any(*SEARCH_ROLES):
        return JSONResponse({"error": "Sin autorización"}, status_code=403)
    record = db.query(DocumentRecord).filter(DocumentRecord.id == record_id).first()
    if not record:
        return JSONResponse({"error": "Registro no encontrado"}, status_code=404)

    subject, body = _build_email_body(record)
    # Gmail compose — abre Gmail en el navegador con asunto y cuerpo pre-llenados.
    # Requiere que el usuario ya tenga sesión iniciada en Gmail.
    # Nota: Gmail admite cuerpo vía URL hasta ~1800 chars; lo truncamos si excede.
    body_encoded = quote(body[:1800])
    gmail_url = (
        "https://mail.google.com/mail/?view=cm&fs=1"
        f"&su={quote(subject)}"
        f"&body={body_encoded}"
    )
    return JSONResponse({
        "gmail_url": gmail_url,
        "pdf_url": f"/pdf/{record_id}",
        "pdf_name": record.nombre_archivo_pdf or "aviso.pdf",
    })


@app.get("/email/{record_id}")
def email_record(record_id: int, request: Request, db: Session = Depends(get_db)):
    """Redirige directamente a Outlook web compose (fallback sin JS)."""
    user = _get_user(request, db)
    if not user or not user.has_any(*SEARCH_ROLES):
        return RedirectResponse("/login", status_code=302)
    record = db.query(DocumentRecord).filter(DocumentRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "Registro no encontrado")

    subject, body = _build_email_body(record)
    gmail_url = (
        "https://mail.google.com/mail/?view=cm&fs=1"
        f"&su={quote(subject)}"
        f"&body={quote(body[:1800])}"
    )
    return RedirectResponse(gmail_url, status_code=302)
